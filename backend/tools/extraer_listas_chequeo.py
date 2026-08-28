#!/usr/bin/env python3
"""Convierte las listas de chequeo oficiales (.xlsx de CompromISO) al catálogo
JSON que consume el backend (`src/main/resources/listas-chequeo/`).

No reescribe ni interpreta el contenido: copia el texto de cada celda tal cual
y solo deriva metadatos verificables (código de formato citado, marca "*" de
"cuando aplique", agrupación por etapa). Lo que el archivo original trae mal o
incompleto se reporta en `advertencias`, no se corrige en silencio.

Uso:
    pip install openpyxl
    python backend/tools/extraer_listas_chequeo.py <carpeta-con-los-xlsx>

Para regenerar el catálogo cuando SENA publique una versión nueva de un
formato: reemplazar el .xlsx en la carpeta de origen y volver a correr esto.
"""

import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

SALIDA = Path(__file__).resolve().parents[1] / "src/main/resources/listas-chequeo"

# Un código de formato institucional: GCCON-F-046, GCCON-AN-001, GJ-F-026,
# GD-F-010, GRF-F-089, GIL-F-029... El separador puede venir con espacios de
# más ("GCCON - F- 056"), así que se normaliza después de capturarlo.
RE_FORMATO = re.compile(r"\b([A-Z]{2,5})\s*-\s*([A-Z]{1,3})\s*-\s*(\d{2,3})\b")
# GTH-074 no sigue el patrón de tres segmentos, pero es un formato citado.
RE_FORMATO_CORTO = re.compile(r"\b(GTH)\s*-\s*(\d{3})\b")

RE_VERSION = re.compile(r"Versi[oó]n:\s*(\S+)", re.IGNORECASE)
RE_CODIGO = re.compile(r"C[oó]digo:\s*(.+)", re.IGNORECASE | re.DOTALL)
RE_VERSION_ARCHIVO = re.compile(r"[_-]?V(\d{2})\b", re.IGNORECASE)

# Formatos de contratación (GCCON): todos comparten el mismo diseño de hoja
# (fila 12 = encabezados, fila 13 = SI/NO, datos desde la fila 14).
LISTAS_GCCON = [
    ("GCCON-F-026-ListadeChequeoContratacionDirectaV04.xlsx", "Contratación Directa", "MODALIDAD_SELECCION"),
    ("GCCON-F-049-ListadeChequeoConcursodemeritosV03.xlsx", "Concurso de méritos", "MODALIDAD_SELECCION"),
    ("GCCON-F-051-ListadeChequeoLicitacionPublicaV02.xlsx", "Propuesta de lista", "MODALIDAD_SELECCION"),
    ("GCCON-F-053-ListadeChequeoMinimaCuantia_V05.xlsx", "Lista de Chequeo", "MODALIDAD_SELECCION"),
    ("GCCON-F-055-ListadeChequeoAbreviadaporSubastaInversaV02.xlsx", "Lista de Chequeo", "MODALIDAD_SELECCION"),
    ("GCCON-F-056-ListadeChequeoSeleccionAbreviadaporMenorCuantiaV02.xlsx", "Lista de Chequeo", "MODALIDAD_SELECCION"),
    ("GCCON-F-052-TerminacionAnticipadayLiquidacionV04.xlsx", "Liquidación y Terminación Antci", "TRAMITE_CONTRACTUAL"),
]

# GRF-F-088 no se agrupa por etapa sino por tipo de pago (una columna por tipo).
# El código de cada tipo es un identificador nuestro; el nombre es el texto
# literal del encabezado de esa columna en el archivo original.
ARCHIVO_PAGO = "GRF-F-088FormatoListadeChequeoDocumentosRequeridosparaelPago(2).xlsx"
CODIGOS_TIPO_PAGO = {
    "C": "ADQUISICION_BIENES",
    "D": "IMPORTACION_BIENES",
    "E": "ADQUISICION_SERVICIOS",
    "F": "OTROS_SERVICIOS_SIN_CONTRATO",
    "G": "SERVICIOS_OBRA_CIVIL",
    "H": "BOLSA_MERCANTIL",
    "I": "HONORARIOS_PRESTACION_SERVICIOS",
    "J": "ANTICIPOS_Y_RECURSOS_EN_ADMINISTRACION",
    "K": "CONVENIOS",
    "L": "CONVENIOS_COOPERACION_INTERNACIONAL",
    "M": "FONDO_VIVIENDA",
    "N": "CONVENIOS_COLUMNA_N",
    "O": "CONVENIOS_COOPERACION_INTERNACIONAL_COLUMNA_O",
    "P": "FONDO_VIVIENDA_COLUMNA_P",
}


def texto(celda):
    valor = celda.value if hasattr(celda, "value") else celda
    if valor is None:
        return ""
    return re.sub(r"[ \t]+", " ", str(valor)).strip()


def normalizar_codigo(bruto):
    return re.sub(r"\s+", "", bruto).upper().replace("\u2013", "-")


def formatos_citados(*textos):
    encontrados = []
    for t in textos:
        for regex in (RE_FORMATO, RE_FORMATO_CORTO):
            for coincidencia in regex.finditer(t.upper()):
                codigo = normalizar_codigo("-".join(coincidencia.groups()))
                if codigo not in encontrados:
                    encontrados.append(codigo)
    return encontrados


def sin_tildes(t):
    return "".join(c for c in unicodedata.normalize("NFD", t) if unicodedata.category(c) != "Mn")


def version_del_archivo(nombre):
    coincidencia = RE_VERSION_ARCHIVO.search(nombre.replace(".xlsx", ""))
    return coincidencia.group(1) if coincidencia else None


def encabezado(hoja, columnas, regex):
    """Busca en las primeras filas la celda "Versión: NN" / "Código: XXX"."""
    for fila in range(1, 6):
        for col in columnas:
            coincidencia = regex.search(texto(hoja["{}{}".format(col, fila)]))
            if coincidencia:
                return coincidencia.group(1).strip()
    return None


def extraer_gccon(ruta, nombre_hoja, tipo):
    libro = openpyxl.load_workbook(ruta, data_only=True)
    hoja = libro[nombre_hoja]
    advertencias = []

    codigo = normalizar_codigo(encabezado(hoja, "HK", RE_CODIGO) or "")
    version = encabezado(hoja, "HK", RE_VERSION) or ""
    version_archivo = version_del_archivo(ruta.name)
    if version_archivo and version_archivo != version:
        advertencias.append(
            "El nombre del archivo dice versión {} pero el encabezado de la hoja dice versión {}. "
            "Se conserva la del encabezado.".format(version_archivo, version)
        )

    for ignorada in [h.title for h in libro.worksheets if h.title not in (nombre_hoja, "INSTRUCTIVO")]:
        advertencias.append(
            'El archivo trae una hoja adicional "{}" que no hace parte del formato publicado; '
            "no se incluyó en el catálogo.".format(ignorada)
        )

    etapas, notas = [], []
    numeros_vistos = set()
    for fila in range(14, hoja.max_row + 1):
        item = hoja["A{}".format(fila)].value
        documento = texto(hoja["B{}".format(fila)])
        observacion = texto(hoja["G{}".format(fila)])

        if isinstance(item, (int, float)) and documento:
            if not etapas:
                # GCCON-F-055 no rotula la primera etapa: la fila del título
                # viene vacía. Se agrupa igual, marcándolo como no rotulado.
                etapas.append({"nombre": "ETAPA PRECONTRACTUAL", "rotuladaEnOrigen": False, "items": []})
                advertencias.append(
                    "La primera etapa no viene rotulada en el formato original; los ítems previos "
                    "a la ETAPA CONTRACTUAL se agruparon como ETAPA PRECONTRACTUAL."
                )
            numero = int(item)
            if numero in numeros_vistos:
                advertencias.append(
                    "El formato original numera dos ítems distintos con el número {}. "
                    "Se conservó la numeración tal como aparece.".format(numero)
                )
            numeros_vistos.add(numero)
            etapas[-1]["items"].append({
                "numero": numero,
                "documento": documento.rstrip("* ").strip(),
                "cuandoAplique": documento.rstrip().endswith("*"),
                "observacion": observacion,
                "formatos": formatos_citados(documento, observacion),
                "tiposPago": [],
            })
            continue

        titulo = texto(hoja["A{}".format(fila)])
        if not titulo:
            continue
        if "ETAPA" in sin_tildes(titulo).upper() and not titulo.startswith("*"):
            etapas.append({"nombre": titulo, "rotuladaEnOrigen": True, "items": []})
        else:
            notas.append(titulo)

    return {
        "codigo": codigo,
        "nombre": texto(hoja["A6"]),
        "version": version,
        "proceso": texto(hoja["A4"]),
        "tipo": tipo,
        "alcance": texto(hoja["D10"]),
        "archivoOrigen": ruta.name,
        "hojaOrigen": nombre_hoja,
        "tiposPago": [],
        "etapas": etapas,
        "notas": notas,
        "advertencias": advertencias,
    }


def extraer_pago(ruta):
    libro = openpyxl.load_workbook(ruta, data_only=True)
    hoja = libro["Formato GRF-F-088"]
    advertencias = []

    codigo = normalizar_codigo(encabezado(hoja, "JM", RE_CODIGO) or "")
    version = encabezado(hoja, "JM", RE_VERSION) or ""

    columnas = [c for c in CODIGOS_TIPO_PAGO if texto(hoja["{}10".format(c)])]
    usadas = {
        c for c in columnas
        for fila in range(11, hoja.max_row + 1)
        if texto(hoja["{}{}".format(c, fila)]).upper() == "X"
    }
    for vacia in [c for c in columnas if c not in usadas]:
        advertencias.append(
            'La columna {} ("{}") no marca ningún documento en el archivo original y repite el '
            "nombre de otra columna; no se incluyó en el catálogo.".format(
                vacia, texto(hoja["{}10".format(vacia)]))
        )
    columnas = [c for c in columnas if c in usadas]

    tipos_pago = [
        {"codigo": CODIGOS_TIPO_PAGO[c], "nombre": texto(hoja["{}10".format(c)])} for c in columnas
    ]

    items = []
    for fila in range(11, hoja.max_row + 1):
        item = hoja["A{}".format(fila)].value
        documento = texto(hoja["B{}".format(fila)])
        if not isinstance(item, (int, float)) or not documento:
            continue
        items.append({
            "numero": int(item),
            "documento": documento.rstrip("* ").strip(),
            "cuandoAplique": documento.rstrip().endswith("*"),
            "observacion": "",
            "formatos": formatos_citados(documento),
            "tiposPago": [
                CODIGOS_TIPO_PAGO[c] for c in columnas
                if texto(hoja["{}{}".format(c, fila)]).upper() == "X"
            ],
        })

    advertencias.append(
        "El formato no agrupa los documentos por etapa sino por tipo de pago; el catálogo los "
        "expone en un único grupo y la aplicabilidad de cada ítem queda en su lista tiposPago."
    )

    return {
        "codigo": codigo,
        "nombre": texto(hoja["A6"]),
        "version": version,
        "proceso": texto(hoja["A4"]),
        "tipo": "TRAMITE_PAGO",
        "alcance": "REGISTRO DE OBLIGACIONES Y TRÁMITE DE PAGO",
        "archivoOrigen": ruta.name,
        "hojaOrigen": "Formato GRF-F-088",
        "tiposPago": tipos_pago,
        "etapas": [{
            "nombre": "DOCUMENTOS REQUERIDOS PARA REGISTRO DE OBLIGACIONES",
            "rotuladaEnOrigen": False,
            "items": items,
        }],
        "notas": [],
        "advertencias": advertencias,
    }


def main():
    if len(sys.argv) != 2:
        print(__doc__)
        return 1
    origen = Path(sys.argv[1])
    SALIDA.mkdir(parents=True, exist_ok=True)

    catalogos = [extraer_gccon(origen / archivo, hoja, tipo) for archivo, hoja, tipo in LISTAS_GCCON]
    catalogos.append(extraer_pago(origen / ARCHIVO_PAGO))

    for catalogo in sorted(catalogos, key=lambda c: c["codigo"]):
        destino = SALIDA / "{}.json".format(catalogo["codigo"])
        destino.write_text(json.dumps(catalogo, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        total = sum(len(e["items"]) for e in catalogo["etapas"])
        print("{:<12} v{:<3} {:>3} ítems  {} etapa(s)  {} advertencia(s)  -> {}".format(
            catalogo["codigo"], catalogo["version"], total,
            len(catalogo["etapas"]), len(catalogo["advertencias"]), destino.name))
    return 0


if __name__ == "__main__":
    sys.exit(main())
